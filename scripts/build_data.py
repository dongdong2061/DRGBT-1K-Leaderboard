from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PRIMARY_SOURCE = ROOT / "DRGBT1k_results.xlsx"
LEGACY_SOURCE = ROOT / "results.xlsx"
METHODS_CONFIG = ROOT / "methods.config.json"
OUTPUT_PATH = ROOT / "data.js"

TRAIN_SEQUENCES = 800
TEST_SEQUENCES = 245

BATCH_LABELS = {
    "第一批方法": "Batch 1",
    "第二批方法": "Batch 2",
}

CATEGORY_LABELS = {
    "全微调 RGBT": "Full Fine-tuning RGBT",
    "部分微调 RGBT": "Partial Fine-tuning RGBT",
    "多模态跟踪 (仅训练RGBT)": "Multimodal Tracking (RGBT-Only Training)",
}

DEFAULT_CATEGORY = "全微调 RGBT"


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def normalize_method_key(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\(.*?\)", "", text)
    return re.sub(r"[^a-z0-9]+", "", text)


def parse_metric(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def extract_year_number(raw_value: object) -> int | None:
    text = str(raw_value or "").strip()
    if not text:
        return None

    matches = re.findall(r"(20\d{2}|\d{2})", text)
    if not matches:
        return None

    token = matches[-1]
    if len(token) == 4:
        return int(token)
    return 2000 + int(token)


def extract_venue(raw_value: object) -> str:
    text = str(raw_value or "").strip()
    if not text:
        return ""
    venue = re.sub(r"[^A-Za-z]+", " ", text).strip()
    return re.sub(r"\s+", " ", venue)


def clean_publication_label(raw_value: object) -> str:
    text = str(raw_value or "").strip()
    if not text:
        return ""

    year_number = extract_year_number(text)
    venue = extract_venue(text)

    if venue and year_number:
        return f"{venue} {year_number}"
    if year_number:
        return str(year_number)
    if venue:
        return venue
    return text


def parse_resource(value: object) -> tuple[str | None, str | None]:
    candidate = str(value or "").strip()
    if not candidate:
        return None, None

    if re.match(r"^(https?:)?//", candidate) or candidate.startswith(("./", "../", "/")):
        return candidate, None
    if candidate.lower().startswith("doi:"):
        return f"https://doi.org/{candidate[4:].strip()}", None
    if re.match(r"^[A-Za-z]:\\", candidate):
        return None, candidate
    if re.match(r"^[A-Za-z0-9._-]+\.[A-Za-z]{2,}", candidate):
        return f"https://{candidate}", None
    return None, candidate


def translate_batch(label: str) -> str:
    return BATCH_LABELS.get(label, label or "Batch")


def translate_category(label: str) -> str:
    return CATEGORY_LABELS.get(label, label or "Uncategorized")


def load_method_config() -> dict[str, dict]:
    if not METHODS_CONFIG.exists():
        return {}

    entries = json.loads(METHODS_CONFIG.read_text(encoding="utf-8"))
    config_map: dict[str, dict] = {}

    for entry in entries:
        key = normalize_method_key(entry.get("id") or entry.get("name"))
        if not key:
            continue

        result_url, result_local = parse_resource(entry.get("result_url") or entry.get("result_path"))
        code_url, code_local = parse_resource(entry.get("code_url") or entry.get("code_path"))
        weight_url, weight_local = parse_resource(entry.get("weight_url") or entry.get("weight_path"))

        config_map[key] = {
            "result": {"url": result_url, "localPath": result_local},
            "code": {"url": code_url, "localPath": code_local},
            "weight": {"url": weight_url, "localPath": weight_local},
        }

    return config_map


def build_resources(paper_url: str | None, method_key: str, config_map: dict[str, dict]) -> list[dict]:
    config = config_map.get(method_key, {})
    result = config.get("result", {})
    code = config.get("code", {})
    weight = config.get("weight", {})

    return [
        {
            "key": "paper",
            "label": "Paper",
            "url": paper_url,
            "localPath": None,
        },
        {
            "key": "result",
            "label": "Tracking Result",
            "url": result.get("url"),
            "localPath": result.get("localPath"),
        },
        {
            "key": "code",
            "label": "Code",
            "url": code.get("url"),
            "localPath": code.get("localPath"),
        },
        {
            "key": "weight",
            "label": "Weight",
            "url": weight.get("url"),
            "localPath": weight.get("localPath"),
        },
    ]


def method_looks_like_section(
    method_name: str,
    year_raw: str,
    pr: float | None,
    npr: float | None,
    sr: float | None,
    paper_title: str,
    paper_url: str | None,
) -> bool:
    if not method_name:
        return False
    if year_raw or paper_title or paper_url:
        return False
    if any(value is not None for value in (pr, npr, sr)):
        return False
    return re.match(r"^[A-Za-z0-9]", method_name) is None


def detect_columns(headers: list[object]) -> dict[str, int | None]:
    columns: dict[str, int | None] = {
        "paper_title": None,
        "method": None,
        "year": None,
        "pr": None,
        "npr": None,
        "sr": None,
        "paper": None,
    }

    for index, raw_header in enumerate(headers, start=1):
        normalized = normalize_header(raw_header)
        if not normalized:
            continue

        if normalized in {"论文名", "papertitle", "title"}:
            columns["paper_title"] = index
        elif normalized in {"method", "方法", "方法名"} or "训练方式" in normalized:
            columns["method"] = index
        elif normalized in {"year", "发表年份", "time"}:
            columns["year"] = index
        elif normalized == "pr":
            columns["pr"] = index
        elif normalized == "npr":
            columns["npr"] = index
        elif normalized == "sr":
            columns["sr"] = index
        elif normalized in {"paper", "paperlink", "论文链接"}:
            columns["paper"] = index

    return columns


def cell_value(ws, row_index: int, column_index: int | None) -> object:
    if not column_index:
        return None
    return ws.cell(row=row_index, column=column_index).value


def cell_link(ws, row_index: int, column_index: int | None) -> str | None:
    if not column_index:
        return None
    hyperlink = ws.cell(row=row_index, column=column_index).hyperlink
    if not hyperlink:
        return None
    return getattr(hyperlink, "target", None) or getattr(hyperlink, "location", None)


def parse_multisheet_workbook(path: Path, config_map: dict[str, dict]) -> dict:
    workbook = load_workbook(path, data_only=True)
    methods: list[dict] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        headers = [ws.cell(row=1, column=index).value for index in range(1, ws.max_column + 1)]
        columns = detect_columns(headers)
        current_category = DEFAULT_CATEGORY

        for row_index in range(2, ws.max_row + 1):
            method_name = str(cell_value(ws, row_index, columns.get("method")) or "").strip()
            if not method_name:
                continue

            paper_title = str(cell_value(ws, row_index, columns.get("paper_title")) or "").strip()
            paper_title = re.sub(r"\s+", " ", paper_title)
            year_raw = str(cell_value(ws, row_index, columns.get("year")) or "").strip()
            pr = parse_metric(cell_value(ws, row_index, columns.get("pr")))
            npr = parse_metric(cell_value(ws, row_index, columns.get("npr")))
            sr = parse_metric(cell_value(ws, row_index, columns.get("sr")))
            paper_value = cell_link(ws, row_index, columns.get("paper")) or cell_value(ws, row_index, columns.get("paper"))
            paper_url, _ = parse_resource(paper_value)

            if method_looks_like_section(method_name, year_raw, pr, npr, sr, paper_title, paper_url):
                current_category = method_name
                continue

            method_key = normalize_method_key(method_name)
            status = "evaluated" if all(value is not None for value in (pr, npr, sr)) else "pending"

            methods.append(
                {
                    "id": f"{translate_batch(sheet_name).lower().replace(' ', '-')}-{method_key or row_index}",
                    "name": method_name,
                    "paperTitle": paper_title,
                    "publication": clean_publication_label(year_raw),
                    "timeRaw": year_raw,
                    "yearNumber": extract_year_number(year_raw),
                    "batchLabel": translate_batch(sheet_name),
                    "categoryLabel": translate_category(current_category),
                    "status": status,
                    "hasMetrics": all(value is not None for value in (pr, npr, sr)),
                    "metrics": {
                        "PR": pr,
                        "NPR": npr,
                        "SR": sr,
                    },
                    "resources": build_resources(paper_url, method_key, config_map),
                }
            )

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "dataset": {
            "name": "DRGBT1K",
            "trainSequences": TRAIN_SEQUENCES,
            "testSequences": TEST_SEQUENCES,
            "metrics": ["PR", "NPR", "SR"],
            "source": path.name,
        },
        "methods": methods,
    }


def parse_legacy_workbook(path: Path, config_map: dict[str, dict]) -> dict:
    workbook = load_workbook(path, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    headers = [normalize_header(ws.cell(row=1, column=index).value) for index in range(1, ws.max_column + 1)]
    column_map = {header: index + 1 for index, header in enumerate(headers) if header}

    required_headers = {"method", "year", "pr", "npr", "sr"}
    missing = required_headers - set(column_map)
    if missing:
        raise ValueError(f"Missing required headers in legacy workbook: {sorted(missing)}")

    methods = []
    for row_index in range(2, ws.max_row + 1):
        method_name = str(cell_value(ws, row_index, column_map["method"]) or "").strip()
        if not method_name:
            continue

        year_raw = str(cell_value(ws, row_index, column_map["year"]) or "").strip()
        pr = parse_metric(cell_value(ws, row_index, column_map["pr"]))
        npr = parse_metric(cell_value(ws, row_index, column_map["npr"]))
        sr = parse_metric(cell_value(ws, row_index, column_map["sr"]))
        paper_value = cell_link(ws, row_index, column_map.get("paper")) or cell_value(ws, row_index, column_map.get("paper"))
        paper_url, _ = parse_resource(paper_value)
        method_key = normalize_method_key(method_name)

        methods.append(
            {
                "id": f"legacy-{method_key or row_index}",
                "name": method_name,
                "paperTitle": "",
                "publication": clean_publication_label(year_raw),
                "timeRaw": year_raw,
                "yearNumber": extract_year_number(year_raw),
                "batchLabel": "Legacy Results",
                "categoryLabel": "Full Fine-tuning RGBT",
                "status": "evaluated" if all(value is not None for value in (pr, npr, sr)) else "pending",
                "hasMetrics": all(value is not None for value in (pr, npr, sr)),
                "metrics": {
                    "PR": pr,
                    "NPR": npr,
                    "SR": sr,
                },
                "resources": build_resources(paper_url, method_key, config_map),
            }
        )

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "dataset": {
            "name": "DRGBT1K",
            "trainSequences": TRAIN_SEQUENCES,
            "testSequences": TEST_SEQUENCES,
            "metrics": ["PR", "NPR", "SR"],
            "source": path.name,
        },
        "methods": methods,
    }


def load_payload() -> dict:
    config_map = load_method_config()
    if PRIMARY_SOURCE.exists():
        return parse_multisheet_workbook(PRIMARY_SOURCE, config_map)
    if LEGACY_SOURCE.exists():
        return parse_legacy_workbook(LEGACY_SOURCE, config_map)
    raise FileNotFoundError("Neither DRGBT1k_results.xlsx nor results.xlsx was found.")


def main() -> None:
    payload = load_payload()
    OUTPUT_PATH.write_text(
        "window.DRGBT_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
